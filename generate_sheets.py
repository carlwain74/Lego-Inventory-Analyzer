"""
This is neat little utility that gets pricing for Lego sets.
"""
import json
import argparse
import sys
import logging
import os
from os import stat
from os.path import exists
import html
from html.parser import HTMLParser
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment,Font,PatternFill
from datetime import datetime


logging.basicConfig(
format='%(asctime)s %(levelname)-8s %(message)s',
level=logging.INFO,
datefmt='%Y-%m-%d %H:%M:%S')

"""
Create workbook
"""
def create_wookbook(xls_filename):
    logging.info(f"Creating workbook: {xls_filename}")
    try:
        if os.path.isfile(xls_filename) and os.access(xls_filename, os.R_OK):
            logging.info(f'Load excel file: {xls_filename}')
            workbook = load_workbook(filename=xls_filename)
        else:
            workbook = Workbook()
            logging.info(workbook.sheetnames)
            workbook.remove(workbook['Sheet'])
    except Exception as exception:
        logging.error(f'Could not load excel file: {xls_filename} -  {str(exception)}')
        sys.exit(1)

    return workbook

"""
Add workbook unless it already exists
"""
def add_worksheet(workbook, item_name):
    # See if the worksheet already exists
    if item_name in workbook.sheetnames:
        worksheet = workbook[item_name]
    else:
        worksheet = workbook.create_sheet(item_name, 0)

        # Start from the first cell. Rows and columns are zero indexed.
        _row = 2
        _col = 2

        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 20

        header_color = "00C0C0C0"
        data = worksheet.cell(row=2, column=2, value="Name")
        data.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        data.alignment = Alignment(horizontal="center", vertical="center")

        data = worksheet.cell(row=3, column=2, value="Category")
        data.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        data.alignment = Alignment(horizontal="center", vertical="center")

        xls_headers = ['Date', 'Avg Price', 'Min Price', 'Max Price', 'Quantity']

        _row = 5
        col_adjust = 0
        for headers in xls_headers:
            #worksheet.write(row, col+col_adjust, headers, header_format)
            data = worksheet.cell(row=_row, column=_col+col_adjust, value=headers)
            data.alignment = Alignment(horizontal="center", vertical="center")
            data.fill = PatternFill(start_color=header_color,
                                    end_color=header_color, fill_type="solid")
            col_adjust += 1

    return worksheet

def create_wookbook_and_sheet(xls_filename):
    workbook = create_wookbook(xls_filename)

    now = datetime.now() # current date and time
    date_stamp = now.strftime("%m_%d_%Y")
    worksheet = workbook.create_sheet('Items_'+date_stamp)

    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    col = 1

    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['F'].width = 20

    header_color = "00C0C0C0"

    xls_headers = ['Item', 'Name', 'Category', 'Avg Price', 'Min Price', 'Max Price', 'Quantity', 'Year']

    _row = 5
    col_adjust = 0
    for headers in xls_headers:
        #worksheet.write(row, col+col_adjust, headers, header_format)
        data = worksheet.cell(row=row, column=col+col_adjust, value=headers)
        data.alignment = Alignment(horizontal="center", vertical="center")
        data.fill = PatternFill(start_color=header_color,
                                end_color=header_color, fill_type="solid")
        col_adjust += 1

    return workbook, worksheet


def generate_single_sheet(sets, workbook, worksheet):
    logging.info('Writing all sets to the same file')
    total = 0
    _row = 1
    _col = 1
    for _set in sets:
        _row += 1

        data = worksheet.cell(row=_row, column=_col, value=_set)
        data.alignment = Alignment(horizontal="center", vertical="center")
        data = worksheet.cell(row=_row, column=_col+1, value=sets[_set]['name'])
        data.alignment = Alignment(horizontal="center", vertical="center")
        data = worksheet.cell(row=_row, column=_col+2, value=sets[_set]['category'])
        data.alignment = Alignment(horizontal="center", vertical="center")
        data = worksheet.cell(row=_row, column=_col+3, value=sets[_set]['current']['avg'])
        data.alignment = Alignment(horizontal="center", vertical="center")
        data = worksheet.cell(row=_row, column=_col+4, value=sets[_set]['current']['min'])
        data.alignment = Alignment(horizontal="center", vertical="center")
        data = worksheet.cell(row=_row, column=_col+5, value=sets[_set]['current']['max'])
        data.alignment = Alignment(horizontal="center", vertical="center")
        data = worksheet.cell(row=_row, column=_col+6, value=sets[_set]['current']['quantity'])
        data.alignment = Alignment(horizontal="center", vertical="center")
        data = worksheet.cell(row=_row, column=_col+7, value=sets[_set]['year'])
        data.alignment = Alignment(horizontal="center", vertical="center")

def generate_multi_sheet(sets, workbook):

    logging.info("Writing sets per sheet`")

    total = 0
    _row = 6
    _col = 2

    now = datetime.now()
    date_stamp = now.strftime("%m-%d-%Y")

    for _set in sets:
        for key in _set:
            worksheet = add_worksheet(workbook, key)
            # Find next available row on column B
            for index in range(6, 1000):
                if worksheet.cell(row=index, column=2).value is None:
                    _row = index
                    logging.debug('Inserting at ros ' + str(_row))
                    break
                else:
                    logging.debug('Row contents: '+
                                  worksheet.cell(row=index, column=2).value)

            session.print_details(_set[key], key)
            total += res[key]['avg']

            data = worksheet.cell(row=2, column=3, value=_set[key]['name'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=3, column=3, value=_set[key]['category'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col, value=date_stamp)
            data = worksheet.cell(row=_row, column=_col+1, value=_set[key]['current']['avg'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+2, value=_set[key]['current']['min'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+3, value=_set[key]['current']['max'])
            data.alignment = Alignment(horizontal="center", vertical="center")
            data = worksheet.cell(row=_row, column=_col+4, value=_set[key]['current']['quantity'])
            data.alignment = Alignment(horizontal="center", vertical="center")

    logging.info("Total: " + str(total) + "USD")

    if 'Summary' in workbook.sheetnames:
        summary = workbook['Summary']
    else:
        summary = workbook.create_sheet("Summary", 0)

        summary.column_dimensions['B'].width = 10
        summary.column_dimensions['C'].width = 20

        header_color = "00C0C0C0"
        data = summary.cell(row=2, column=2, value="Date")
        data.fill = PatternFill(start_color=header_color,
                                end_color=header_color, fill_type="solid")
        data.alignment = Alignment(horizontal="center", vertical="center")

        data = summary.cell(row=2, column=3, value="Total")
        data.fill = PatternFill(start_color=header_color,
                                end_color=header_color, fill_type="solid")
        data.alignment = Alignment(horizontal="center", vertical="center")

    for index in range(3, 1000):
        if summary.cell(row=index, column=2).value is None:
            _srow = index
            logging.debug('Inserting at ros ' + str(_srow))
            break
        else:
            logging.debug('Row contents: '+summary.cell(row=index, column=2).value)

    data = summary.cell(row=_srow, column=2, value=date_stamp)
    data.alignment = Alignment(horizontal="center", vertical="center")
    data = summary.cell(row=_srow, column=3, value=total)
    data.alignment = Alignment(horizontal="center", vertical="center")

def test_config(config_file = 'config.ini'):
    session = create_api_session(config_file)
    res = getDetails(session, "75105-1")

    if res:
        return True
    else:
        return False

if __name__ == '__main__':
    sheet_handler("71016-1", "", False, False)
